import os
import json
import time
import requests
import io
import traceback
import re
import logging
import uuid
from datetime import datetime
from threading import Thread
from flask import Flask, request, jsonify, render_template, send_file, session
from werkzeug.utils import secure_filename
import openpyxl
from openpyxl.styles import PatternFill
from google import genai                          # <-- new package
from invoice_filter import filter_invoice_pages   # keep only invoice pages

# --- Setup Logging ---
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'sajbbnffnfkqnjf7qw68767qw62')
app.config['MAX_CONTENT_LENGTH'] = 25 * 1024 * 1024

# --- Allowed file extensions ---
ALLOWED_EXTENSIONS = {'pdf', 'png', 'jpg', 'jpeg'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# --- API Configuration ---
NANONETS_API_KEY = os.environ.get("NANONETS_API_KEY")
if not NANONETS_API_KEY:
    raise ValueError("NANONETS_API_KEY environment variable not set")

# Google AI (Gemini) – new package
GOOGLE_AI_API_KEY = os.environ.get("GOOGLE_AI_API_KEY")
genai_client = None
if GOOGLE_AI_API_KEY:
    genai_client = genai.Client(api_key=GOOGLE_AI_API_KEY)
else:
    logger.warning("GOOGLE_AI_API_KEY not set – summaries will be empty")

BATCH_ENDPOINT = "https://extraction-api.nanonets.com/api/v1/extract/batch"
RESULTS_ENDPOINT = "https://extraction-api.nanonets.com/api/v1/extract/results/{}"

EXTRACTION_SCHEMA = {
    "type": "object",
    "properties": {
        "supplier_name": {"type": "string"},
        "invoice_id": {"type": "string"},
        "invoice_date": {"type": "string"},
        "payment_term": {"type": "string"},
        "line_items": {
            "type": "array",
            "items": {
                "type": "object",
                "properties": {
                    "description": {"type": "string"},
                    "quantity": {"type": "number"},
                    "unit_price": {"type": "number"},
                    "discount": {"type": "string"},
                    "tax": {"type": "string"}
                }
            }
        }
    }
}

# In-memory job store
jobs = {}

# ------------------------------------------------------------
# Helper Functions (unchanged)
# ------------------------------------------------------------
def standardize_date(date_str):
    if not date_str or not isinstance(date_str, str):
        return ''
    date_str = date_str.strip()
    formats = [
        "%d/%b/%Y", "%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d.%m.%Y",
        "%d/%m/%y", "%d-%m-%y", "%d.%m.%y", "%d %b %Y", "%d %B %Y",
        "%b %d, %Y", "%B %d, %Y"
    ]
    for fmt in formats:
        try:
            dt = datetime.strptime(date_str, fmt)
            if dt.year < 100:
                dt = dt.replace(year=2000 + dt.year if dt.year < 50 else 1900 + dt.year)
            return dt.strftime("%d/%m/%Y")
        except ValueError:
            continue
    pattern = r'(\d{1,2})[/.-](\d{1,2})[/.-](\d{2,4})'
    match = re.search(pattern, date_str)
    if match:
        day, month, year = match.groups()
        try:
            day, month, year = int(day), int(month), int(year)
            if year < 100:
                year = 2000 + year if year < 50 else 1900 + year
            dt = datetime(year, month, day)
            return dt.strftime("%d/%m/%Y")
        except ValueError:
            pass
    return date_str

def standardize_payment_term(term_str):
    if not term_str or not isinstance(term_str, str):
        return "COD", True
    term_str = term_str.upper()
    numbers = re.findall(r'\d+', term_str)
    if numbers:
        num = int(numbers[0])
        mapping = {7: "NET7", 14: "NET14", 20: "NET20", 25: "NET25", 30: "NET30", 60: "NET60"}
        if num in mapping:
            return mapping[num], False
    return "COD", True

def to_number(value):
    if value is None:
        return 0
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        cleaned = re.sub(r'[^\d.\-]', '', value.replace('%', '').strip())
        if cleaned == '':
            return 0
        try:
            return float(cleaned)
        except ValueError:
            return 0
    return 0

def clean_discount_tax(value):
    if value is None:
        return "0"
    if isinstance(value, (int, float)):
        return str(value)
    if isinstance(value, str):
        stripped = value.strip()
        return stripped if stripped else "0"
    return "0"

def poll_and_store_result(job_id, record_id, filename, headers):
    max_attempts = 120
    interval = 5
    for _ in range(max_attempts):
        try:
            resp = requests.get(
                RESULTS_ENDPOINT.format(record_id),
                headers=headers,
                timeout=10
            )
            resp.raise_for_status()
            data = resp.json()
            status = data.get("status")
            if status == "completed":
                json_result = data.get("result", {}).get("json", {})
                extracted = json_result.get("content", {})
                metadata = json_result.get("metadata", {})
                confidence_scores = metadata.get("confidence_score", {})

                extracted['invoice_date'] = standardize_date(extracted.get('invoice_date'))

                field_confidences = {}
                for field in ["supplier_name", "invoice_id", "invoice_date", "payment_term"]:
                    field_confidences[field] = confidence_scores.get(field)

                line_items_conf = []
                line_items = extracted.get("line_items", [])
                line_items_conf_raw = confidence_scores.get("line_items", {})
                ref_keys = list(line_items_conf_raw.keys())
                for i, item in enumerate(line_items):
                    item_conf = {}
                    if i < len(ref_keys):
                        ref_key = ref_keys[i]
                        item_conf_data = line_items_conf_raw.get(ref_key, {})
                        for subfield in ["description", "quantity", "unit_price", "discount", "tax"]:
                            item_conf[subfield] = item_conf_data.get(subfield)
                    else:
                        for subfield in ["description", "quantity", "unit_price", "discount", "tax"]:
                            item_conf[subfield] = None
                    line_items_conf.append(item_conf)

                all_confidences = list(field_confidences.values())
                for ic in line_items_conf:
                    all_confidences.extend(ic.values())
                valid_confs = [c for c in all_confidences if c is not None]
                avg_confidence = sum(valid_confs) / len(valid_confs) if valid_confs else None

                tags_list = jobs.get(job_id, {}).get('tags', [])

                jobs[job_id] = {
                    "filename": filename,
                    "status": "completed",
                    "result": {
                        "file": filename,
                        "extracted_data": extracted,
                        "field_confidences": field_confidences,
                        "line_items_confidences": line_items_conf,
                        "average_confidence": avg_confidence,
                        "tags": tags_list,
                        "summary": None,
                        "include_summary": True
                    }
                }
                return
            elif status == "failed":
                jobs[job_id] = {
                    "filename": filename,
                    "status": "failed",
                    "error": data.get("message", "Processing failed")
                }
                return
            time.sleep(interval)
        except Exception as e:
            logger.warning(f"Polling error for {job_id}: {e}")
            time.sleep(interval)

    jobs[job_id] = {
        "filename": filename,
        "status": "timeout",
        "error": "Processing timed out"
    }

# ------------------------------------------------------------
# Google AI Summary Helper (unchanged)
# ------------------------------------------------------------
def generate_items_summary(line_items):
    if genai_client is None:
        return ""

    descriptions = [item.get("description", "").strip() for item in line_items if item.get("description")]
    if not descriptions:
        return "No items to summarise."

    items_text = "\n".join(descriptions)
    prompt = "Summarize this pizza shop invoice in one short sentence (about 10 words), categorizing items by purpose such as pizza ingredients, drinks, utensils, packaging, cleaning supplies, or other shop essentials. Do not include “Invoice summary.” \n\n" + items_text

    try:
        response = genai_client.models.generate_content(
            model="gemini-2.5-flash",
            contents=prompt
        )
        summary = response.text.strip()
        words = summary.split()
        if len(words) > 25:
            summary = " ".join(words[:25]) + "..."
        return summary
    except Exception as e:
        logger.error(f"AI summary generation failed: {e}")
        return ""

# ------------------------------------------------------------
# Routes
# ------------------------------------------------------------
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/health')
def health():
    return jsonify({"status": "ok"}), 200

@app.route('/extract', methods=['POST'])
def submit_extraction():
    try:
        if 'files' not in request.files:
            return jsonify({"error": "No files uploaded"}), 400

        files = request.files.getlist('files')
        if not files:
            return jsonify({"error": "Empty file list"}), 400

        for f in files:
            if f.filename == '':
                return jsonify({"error": "Empty filename detected"}), 400
            if not allowed_file(f.filename):
                return jsonify({"error": f"File type not allowed: {f.filename}. Only PDF, PNG, JPG, JPEG are allowed."}), 400

        tags_mapping = {}
        if 'tags_mapping' in request.form:
            try:
                tags_mapping = json.loads(request.form['tags_mapping'])
            except json.JSONDecodeError:
                logger.warning("Invalid tags_mapping JSON, ignoring")

        multipart_data = {
            'output_format': 'json',
            'json_options': json.dumps(EXTRACTION_SCHEMA),
            'include_metadata': 'confidence_score'
        }

        file_tuples = []
        file_info = []

        for f in files:
            if f.filename == '':
                continue
            filename = secure_filename(f.filename)
            file_content = f.read()

            # Keep only the invoice page(s) before sending to Nanonets. The
            # uploaded file is a merged pack (Invoice + PR + GRN) and Nanonets
            # would otherwise pick the PR/GRN number or mix in their line
            # items. Filtering down to the invoice page fixes both problems.
            try:
                file_content, filter_info = filter_invoice_pages(
                    file_content, filename=filename, genai_client=genai_client
                )
                logger.info(
                    f"Invoice page filter [{filename}]: {filter_info['reason']} "
                    f"| labels={filter_info['labels']} "
                    f"| kept={filter_info['kept_pages']}"
                )
            except Exception as e:
                logger.warning(
                    f"Invoice page filtering failed for {filename}: {e}; "
                    f"sending original file unchanged"
                )

            file_tuples.append(('files', (filename, file_content, f.content_type)))
            tags = tags_mapping.get(f.filename, [])
            file_info.append((filename, tags))

        if not file_tuples:
            return jsonify({"error": "No valid files"}), 400

        headers = {"Authorization": f"Bearer {NANONETS_API_KEY}"}
        batch_resp = requests.post(
            BATCH_ENDPOINT,
            headers=headers,
            files=file_tuples,
            data=multipart_data,
            timeout=100
        )
        batch_resp.raise_for_status()
        batch_data = batch_resp.json()
        records = batch_data.get("records", [])

        job_ids = []
        for idx, record in enumerate(records):
            record_id = record.get("record_id")
            original_filename, tags = file_info[idx] if idx < len(file_info) else (f"file_{idx}.pdf", [])

            if not record_id:
                continue

            job_id = str(uuid.uuid4())
            jobs[job_id] = {
                "record_id": record_id,
                "filename": original_filename,
                "status": "processing",
                "result": None,
                "tags": tags
            }
            job_ids.append(job_id)

            Thread(target=poll_and_store_result, args=(job_id, record_id, original_filename, headers), daemon=True).start()

        return jsonify({"job_ids": job_ids})

    except Exception as e:
        logger.error(f"Submission failed: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500

@app.route('/status/<job_id>')
def get_status(job_id):
    job = jobs.get(job_id)
    if not job:
        return jsonify({"error": "Job not found"}), 404
    return jsonify(job)

@app.route('/results/batch', methods=['POST'])
def get_batch_results():
    data = request.get_json()
    job_ids = data.get('job_ids', [])
    results = []
    for jid in job_ids:
        job = jobs.get(jid)
        if job and job.get('status') == 'completed':
            results.append(job['result'])
        elif job and job.get('status') == 'failed':
            results.append({"file": job['filename'], "error": job.get('error')})
        elif job and job.get('status') == 'processing':
            results.append({"file": job['filename'], "status": "processing"})
    return jsonify({"results": results})

# NEW: Generate summaries for existing results
@app.route('/generate-summaries', methods=['POST'])
def generate_summaries():
    try:
        data = request.get_json()
        invoices = data.get('results', [])
        for inv in invoices:
            if inv.get('error') or inv.get('status') == 'processing':
                continue
            line_items = inv.get('extracted_data', {}).get('line_items', [])
            summary = generate_items_summary(line_items)
            inv['summary'] = summary
            inv['include_summary'] = inv.get('include_summary', True)  # keep existing or default
        return jsonify({"results": invoices})
    except Exception as e:
        logger.error(f"Generate summaries failed: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500

@app.route('/export', methods=['POST'])
def export_excel():
    try:
        data = request.get_json(silent=True)
        if data and 'results' in data:
            results = data['results']
        elif 'extraction_results' in session:
            results = json.loads(session['extraction_results'])
        else:
            return jsonify({"error": "No extraction results available"}), 400

        template_path = os.path.join(os.path.dirname(__file__), 'purchase_bills.xlsx')
        if not os.path.exists(template_path):
            logger.error(f"Template file not found at {template_path}")
            return jsonify({"error": "Template file purchase_bills.xlsx not found on server"}), 500

        wb = openpyxl.load_workbook(template_path)
        ws = wb.active

        start_row = 31
        current_row = start_row

        yellow_fill_conf = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        red_fill_conf = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        orange_fill_cod = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')

        def get_fill_for_confidence(conf):
            if conf is None:
                return None
            if conf < 90:
                return red_fill_conf
            elif 90 <= conf < 95:
                return yellow_fill_conf
            else:
                return None

        def safe_set_fill(cell, fill):
            if fill is not None:
                try:
                    cell.fill = fill
                except Exception as e:
                    logger.warning(f"Could not set fill on {cell.coordinate}: {e}")

        for invoice in results:
            extracted = invoice.get('extracted_data', {})
            if not extracted:
                continue

            supplier = extracted.get('supplier_name', '')
            invoice_date = extracted.get('invoice_date', '')
            invoice_id = extracted.get('invoice_id', '')
            line_items = extracted.get('line_items', [])

            field_conf = invoice.get('field_confidences', {})
            line_items_conf = invoice.get('line_items_confidences', [])

            raw_payment_term = extracted.get('payment_term', '')
            payment_term, cod_highlight = standardize_payment_term(raw_payment_term)

            tags_list = invoice.get('tags', [])
            tags_str = ', '.join(tags_list) if tags_list else ''

            # Use pre-generated summary and the include flag
            summary = invoice.get('summary', '')
            include_summary = invoice.get('include_summary', True)

            first_item = True
            for i, item in enumerate(line_items):
                if first_item:
                    cell_supplier = ws.cell(row=current_row, column=2)
                    cell_supplier.value = supplier
                    safe_set_fill(cell_supplier, get_fill_for_confidence(field_conf.get('supplier_name')))

                    cell_invoice_id = ws.cell(row=current_row, column=3)
                    cell_invoice_id.value = invoice_id
                    safe_set_fill(cell_invoice_id, get_fill_for_confidence(field_conf.get('invoice_id')))

                    cell_date = ws.cell(row=current_row, column=5)
                    cell_date.value = invoice_date
                    safe_set_fill(cell_date, get_fill_for_confidence(field_conf.get('invoice_date')))

                    cell_payment = ws.cell(row=current_row, column=6)
                    cell_payment.value = payment_term
                    if cod_highlight:
                        safe_set_fill(cell_payment, orange_fill_cod)
                    else:
                        safe_set_fill(cell_payment, get_fill_for_confidence(field_conf.get('payment_term')))

                    ws.cell(row=current_row, column=7).value = ''
                    ws.cell(row=current_row, column=8).value = 'MYR'

                    cell_tags = ws.cell(row=current_row, column=10)
                    cell_tags.value = tags_str

                    # Only write summary if include_summary is True
                    if include_summary:
                        cell_summary = ws.cell(row=current_row, column=11)
                        cell_summary.value = summary
                    else:
                        # Ensure cell is empty if not included
                        ws.cell(row=current_row, column=11).value = None

                    first_item = False

                item_conf = line_items_conf[i] if i < len(line_items_conf) else {}

                cell_desc = ws.cell(row=current_row, column=14)
                cell_desc.value = item.get('description', '')
                safe_set_fill(cell_desc, get_fill_for_confidence(item_conf.get('description')))

                qty = item.get('quantity')
                if qty is not None:
                    cell_qty = ws.cell(row=current_row, column=15)
                    cell_qty.value = qty
                    safe_set_fill(cell_qty, get_fill_for_confidence(item_conf.get('quantity')))

                unit_price = item.get('unit_price')
                if unit_price is not None:
                    cell_price = ws.cell(row=current_row, column=18)
                    cell_price.value = unit_price
                    safe_set_fill(cell_price, get_fill_for_confidence(item_conf.get('unit_price')))

                discount_raw = item.get('discount')
                discount_value = clean_discount_tax(discount_raw)
                cell_disc = ws.cell(row=current_row, column=19)
                cell_disc.value = discount_value
                safe_set_fill(cell_disc, get_fill_for_confidence(item_conf.get('discount')))

                tax_raw = item.get('tax')
                tax_value = clean_discount_tax(tax_raw)
                cell_tax = ws.cell(row=current_row, column=20)
                cell_tax.value = tax_value
                safe_set_fill(cell_tax, get_fill_for_confidence(item_conf.get('tax')))

                current_row += 1

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='extracted_invoices.xlsx'
        )

    except Exception as e:
        logger.error(f"Export failed: {e}", exc_info=True)
        return jsonify({"error": f"Export failed: {str(e)}"}), 500

@app.errorhandler(Exception)
def handle_exception(e):
    logger.error(f"Unhandled exception: {e}", exc_info=True)
    return jsonify({"error": str(e), "type": type(e).__name__}), 500

if __name__ == '__main__':
    port = int(os.environ.get("PORT", 5000))
    app.run(host='0.0.0.0', port=port, debug=False)